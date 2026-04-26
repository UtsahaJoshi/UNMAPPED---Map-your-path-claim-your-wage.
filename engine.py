import openpyxl
import json
import sys

def load_data():
    # Load ISCO Data
    isco_wb = openpyxl.load_workbook('ISCO-08 EN Structure and definitions.xlsx')
    isco_ws = isco_wb.active
    isco_data = []
    
    # Skip header
    for row in isco_ws.iter_rows(min_row=2):
        isco_data.append({
            'code': str(row[1].value),
            'title': str(row[2].value),
            'definition': str(row[3].value),
            'tasks': str(row[4].value),
            'included': str(row[5].value),
            'excluded': str(row[6].value)
        })
        
    # Load Wage Data
    wage_wb = openpyxl.load_workbook('Nepal_Earnings_Occupation_FY2024_25.xlsx')
    wage_ws = wage_wb.active
    wage_data = {}
    
    # Find start of data (looking for rows with numeric values in col 2 or similar)
    for row in wage_ws.iter_rows(min_row=4):
        title = row[0].value
        wage = row[1].value # Assuming col 2 is the wage
        if title and wage:
            wage_data[str(title).strip().lower()] = wage
            
    return isco_data, wage_data

def find_match(query, isco_data):
    query = query.lower()
    matches = []
    
    for item in isco_data:
        score = 0
        if query in item['title'].lower(): score += 10
        if query in item['included'].lower(): score += 5
        if query in item['definition'].lower(): score += 2
        
        # SMART MAPPING: Look at excluded columns to help identify transitions
        if query in item['excluded'].lower():
            # If excluded from a group, it might belong to THIS group's parent or a specific neighbor
            score += 1 
            
        if score > 0:
            matches.append((score, item))
            
    matches.sort(key=lambda x: x[0], reverse=True)
    return [m[1] for m in matches[:5]]

if __name__ == "__main__":
    if len(sys.argv) > 1:
        query = sys.argv[1]
        isco, wages = load_data()
        results = find_match(query, isco)
        print(json.dumps(results, indent=2))
